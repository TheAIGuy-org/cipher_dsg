from __future__ import annotations

import os

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from typing import List

from dotenv import load_dotenv
from llm.azure_client import get_llm_client
from updater.prompt import TableUpdatePrompts

load_dotenv()


class OutputFormat(BaseModel):
    rows:    List[List[str]]
    headers: List[str]


def split_table_content(table_str: str) -> tuple[str, list[str], str]:
    """
    Split generated_text into three parts around the pipe-delimited table.

    Returns:
        prose_before : text before the first pipe row
        pipe_rows    : list of raw pipe-delimited row strings
        prose_after  : text after the last pipe row (totals line, summary sentence)
    """
    lines      = table_str.split("\n")
    first_pipe = None
    last_pipe  = None

    for i, line in enumerate(lines):
        s = line.strip()
        if s.startswith("|") and s.endswith("|") and s.count("|") >= 2:
            if first_pipe is None:
                first_pipe = i
            last_pipe = i

    if first_pipe is None:
        return table_str.strip(), [], ""

    prose_before = "\n\n".join(l.strip() for l in lines[:first_pipe]    if l.strip())
    prose_after  = "\n\n".join(l.strip() for l in lines[last_pipe + 1:] if l.strip())

    # Extract pipe rows, joining wrapped lines onto the previous row
    pipe_rows = []
    pending   = ""
    for line in lines[first_pipe:last_pipe + 1]:
        s = line.strip()
        if not s:
            continue
        if s.startswith("|"):
            if pending:
                pipe_rows.append(pending.strip())
            pending = s
        else:
            pending += " " + s
    if pending:
        pipe_rows.append(pending.strip())

    # Remove separator rows (---|--- style)
    pipe_rows = [
        r for r in pipe_rows
        if r.replace("|", "").strip() and not all(c in "- " for c in r.replace("|", "").strip())
    ]

    return prose_before, pipe_rows, prose_after


def get_table_name(folder_path: str, keywords: list[str]) -> str | None:
    """
    Score each xlsx in folder_path against keywords and return the best match.

    Scoring: +5 per keyword in header row, +1 per keyword in any row,
    +2 if all rows have equal column count, -5 if fewer than 2 rows.
    """
    best_file  = None
    best_score = -1

    for file in os.listdir(folder_path):
        if not file.endswith(".xlsx"):
            continue
        wb   = load_workbook(os.path.join(folder_path, file))
        ws   = wb.active
        data = [
            [str(c).lower() if c else "" for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        if not data:
            continue

        score       = 0
        header_text = " ".join(data[0])
        for kw in keywords:
            kw_l = kw.lower()
            if kw_l in header_text:
                score += 5
            for row in data:
                if kw_l in " ".join(row):
                    score += 1

        col_counts = [len(r) for r in data]
        if len(set(col_counts)) == 1:
            score += 2
        if len(data) < 2:
            score -= 5

        if score > best_score:
            best_score = score
            best_file  = file

    print(f"Best matching file: {best_file}")
    return best_file


def llm_parse_rows(pipe_rows: list[str], folder_path: str) -> tuple[list, list, str | None]:
    """
    Send pipe rows to the LLM and return structured (data_rows, headers, matched_file).

    Calls get_table_name using the LLM-identified headers to find the target xlsx.
    """
    print("Using LLM to parse rows into structured data")

    response = get_llm_client().ask(
        prompt          = TableUpdatePrompts.USER_PROMPT.format(rows=str(pipe_rows)),
        system_prompt   = TableUpdatePrompts.SYSTEM_PROMPT,
        response_format = "json_object",
        temperature     = 0,
    )

    if not response.success:
        raise Exception(f"LLM call failed: {response.error}")

    structured = OutputFormat(**response.content)
    FILE       = get_table_name(folder_path, structured.headers)
    return structured.rows, structured.headers, FILE


def table_to_excel_llm(table_str: str, folder_path: str) -> tuple[str, str]:
    """
    Parse table_str, write the table data to the matching xlsx file,
    and return (prose_before, prose_after) for the caller to handle.

    prose_before is sent to the LLM as a prose UPDATE.
    prose_after is injected directly after the table in the struct.
    """
    print("Making necessary changes to the Excel file")

    prose_before, pipe_rows, prose_after = split_table_content(table_str)

    if not pipe_rows:
        print("No pipe rows found — skipping Excel update")
        return prose_before, prose_after

    data, headers, FILE = llm_parse_rows(pipe_rows, folder_path)

    if not FILE:
        raise ValueError(f"No matching Excel table found in: {folder_path}")

    file_path = os.path.join(folder_path, FILE)
    print(f"Overwriting: {file_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Table"
    ws.append(headers)
    for item in data:
        row_values = list(item)
        while len(row_values) < len(headers):
            row_values.append("")
        ws.append(row_values[:len(headers)])

    wb.save(file_path)
    print(f"✅ Excel overwritten: {file_path}")

    return prose_before, prose_after